from flask import render_template, flash, redirect, url_for, request
from application import app, db, file_manager
from application.forms import LoginForm, RegistrationForm
from flask_login import current_user, login_user, logout_user, login_required
from application.models import User
from werkzeug.urls import url_parse
from datetime import datetime, date
from shutil import copyfile
import os.path
import pandas as pd
import xlsxwriter
import xlrd
import openpyxl
import json

brand = 'Thien\'s Website'
now = datetime.utcnow()


@app.route('/index', methods=['GET', 'POST'])
@app.route('/', methods=['GET', 'POST'])
def index():
    if current_user.is_authenticated:
        return render_template('index.html', brand=brand)

    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            flash('Invalid username or password')
            return render_template('index.html', brand=brand, form=form)
        login_user(user, remember=form.remember_me.data)
        next_page = request.args.get('next')
        if not next_page or url_parse(next_page).netloc != '':
            next_page = url_for('index')
        return redirect(next_page)
    return render_template('index.html', brand=brand, form=form)

# @app.route('/login', methods=['GET', 'POST'])
# def login():
#     if current_user.is_authenticated:
#         return redirect(url_for('index'))
#     form = LoginForm()

#     if form.validate_on_submit():
#         user = User.query.filter_by(username=form.username.data).first()
#         if user is None or not user.check_password(form.password.data):
#             flash('Invalid username or password')
#             return redirect(url_for('login'))
#         login_user(user, remember=form.remember_me.data)
#         # next_page = request.args.get('next')
#         # if not next_page or url_parse(next_page).netloc != '':
#         #     next_page = url_for('index')
#         # return redirect(url_for('next_page'))
#         return redirect(url_for('index'))
#     return render_template('login.html', title='Login', form=form)


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = RegistrationForm()
    if form.validate_on_submit():
        user = User(username=form.username.data, email=form.email.data,
                    fullname=form.fullname.data)
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash('Congratulations, you are now a registered user!')
        return redirect(url_for('index'))
    return render_template('register.html', brand=brand, form=form)


@app.route('/log_out')
def log_out():
    logout_user()
    return redirect(url_for('index'))


@app.route('/charts')
@login_required
def charts():
    return render_template('charts.html', brand=brand, title='Charts')


@app.route('/<uname>')
def profile(uname):
    user = User.query.filter_by(username=uname).first_or_404()
    posts = [
        {'author': uname, 'body': 'Bài viết #1'},
        {'author': uname, 'body': 'Bài viết #2'}
    ]
    return render_template('profile.html', brand=brand, title='Profile',
                           user=user, posts=posts)


@app.before_request
def before_request():
    if current_user.is_authenticated:
        current_user.last_seen = datetime.utcnow()
        db.session.commit()


@app.route('/diagram')
@login_required
def diagram():
    filename = 'test'
    path = app.static_folder + '/excel/' + filename

    data = file_manager.FileManagement()
    data.writeTXTFile(path + '.xlsx', 'Sheet',
                      path.replace('/excel/', '/txt/') + '.txt')
    data.readExcelFile(path + '.xlsx', 'Sheet')
    data_json = data.convertToJSON()
    data.writeJSONFile(path + '.json', data_json)

    return render_template('diagram.html', brand=brand, title='Diagram')


@app.route('/demo-app', methods=['GET', 'POST'])
def demo_app():
    # Get data on server
    path = app.static_folder + '/excel/'
    message = ''

    model_data = pd.read_excel(path + 'Database.xlsx', 'Model')
    dict_model = model_data.to_dict()
    department = pd.read_excel(path + 'Database.xlsx', 'Department')
    dict_dept = department.to_dict()
    line_data = pd.read_excel(path + 'Database.xlsx', 'Line')
    dict_line = line_data.to_dict()

    # Get data Log file
    dict_log = ''
    selected_date = date.today().strftime('%Y-%m-%d')
    selected_log = []  # Log file data

    selected_arm_a = []
    selected_arm_b = []
    selected_arm_c = []
    selected_arm_d = []
    selected_arm_e = []
    selected_arm_f = []
    selected_arm_j = []

    selected_ass_a = []
    selected_ass_b = []
    selected_ass_c = []
    selected_ass_d = []
    selected_ass_e = []
    selected_ass_f = []
    selected_ass_j = []

    if os.path.isfile(path + 'Log.xlsx'):
        log_data = pd.read_excel(path + 'Log.xlsx', 'Sheet1')
        dict_log = log_data.to_dict()

        # Get data of current day in Log file
        for item in dict_log['Date']:
            if dict_log['Date'][item] == selected_date:
                obj = {}
                obj['Date'] = dict_log['Date'][item]
                obj['Time'] = dict_log['Time'][item]
                obj['Department'] = dict_log['Department'][item]
                obj['Line'] = dict_log['Line'][item]
                obj['Model'] = dict_log['Model'][item]
                obj['Plan'] = dict_log['Plan'][item]
                obj['Version'] = dict_log['Version'][item]
                obj['File Name'] = dict_log['File Name'][item]
                selected_log.append(obj)

        # Get data of current day in Plan file
        for item in selected_log:
            # Copy file data and process
            src = path + item['Department'] + '/' \
                + item['Line'] + '/' \
                + item['File Name']
            dst = path + 'ViewData/' + item['Department'] + '/' \
                + item['Line'] + '/' \
                + item['File Name']
            copyfile(src, dst)
            cur_path = dst

            if os.path.isfile(cur_path):
                plan_df = pd.read_excel(cur_path, 'Sheet1')
                last_row = len(plan_df) + 1
                obj = {}  # Object save last row data

                # Write Plan file
                wb = openpyxl.load_workbook(cur_path)
                ws = wb.active

                if last_row > 2:  # Detect new data
                    for i in range(3, last_row + 1):
                        # Copy same value
                        ws['I' + str(i)] = ws['I2'].value  # Date
                        ws['J' + str(i)] = ws['J2'].value  # Department
                        ws['K' + str(i)] = ws['K2'].value  # Line
                        ws['L' + str(i)] = ws['L2'].value  # Model
                        ws['M' + str(i)] = ws['M2'].value  # Version
                        ws['N' + str(i)] = ws['N2'].value  # Q'ty Plan
                        ws['P' + str(i)] = ws['P2'].value  # ST Plan
                        # % Finish
                        cur_qty_actual = int(ws['B' + str(i)].value)
                        cur_qty_plan = int(ws['N' + str(i)].value)
                        ws['O' + str(i)].value = cur_qty_actual / \
                            cur_qty_plan * 100
                        # ST Plan
                        begin_time = datetime.strptime(
                            ws['D3'].value, '%H:%M:%S')
                        cur_time = datetime.strptime(
                            ws['D' + str(i)].value, '%H:%M:%S')
                        delta_second = (cur_time - begin_time).total_seconds()
                        st_actual = (delta_second +
                                     float(ws['P2'].value)) / cur_qty_actual
                        ws['Q' + str(i)].value = st_actual

                # Get data last line of Plan file
                obj = {}
                obj['Start/Stop'] = ws['A' + str(last_row)].value
                obj['Q\'ty Actual'] = ws['B' + str(last_row)].value
                obj['Pause/Continue'] = ws['C' + str(last_row)].value
                obj['Time'] = ws['D' + str(last_row)].value
                obj['Machine'] = ws['E' + str(last_row)].value
                obj['Material'] = ws['F' + str(last_row)].value
                obj['Quality'] = ws['G' + str(last_row)].value
                obj['Other'] = ws['H' + str(last_row)].value
                obj['Date'] = ws['I' + str(last_row)].value
                obj['Department'] = ws['J' + str(last_row)].value
                obj['Line'] = ws['K' + str(last_row)].value
                obj['Model'] = ws['L' + str(last_row)].value
                obj['Version'] = ws['M' + str(last_row)].value
                obj['Q\'ty Plan'] = ws['N' + str(last_row)].value
                obj['% Plan'] = round(ws['O' + str(last_row)].value, 1)
                obj['ST Plan'] = round(ws['P' + str(last_row)].value, 2)
                obj['ST Actual'] = round(ws['Q' + str(last_row)].value, 2)

                wb.save(cur_path)

                if obj['Department'] == 'ARM' and obj['Line'] == 'A':
                    selected_arm_a.append(obj)
                if obj['Department'] == 'ARM' and obj['Line'] == 'B':
                    selected_arm_b.append(obj)
                if obj['Department'] == 'ARM' and obj['Line'] == 'C':
                    selected_arm_c.append(obj)
                if obj['Department'] == 'ARM' and obj['Line'] == 'D':
                    selected_arm_d.append(obj)
                if obj['Department'] == 'ARM' and obj['Line'] == 'E':
                    selected_arm_e.append(obj)
                if obj['Department'] == 'ARM' and obj['Line'] == 'F':
                    selected_arm_f.append(obj)
                if obj['Department'] == 'ARM' and obj['Line'] == 'J':
                    selected_arm_j.append(obj)

                if obj['Department'] == 'ASS' and obj['Line'] == 'A':
                    selected_ass_a.append(obj)
                if obj['Department'] == 'ASS' and obj['Line'] == 'B':
                    selected_ass_b.append(obj)
                if obj['Department'] == 'ASS' and obj['Line'] == 'C':
                    selected_ass_c.append(obj)
                if obj['Department'] == 'ASS' and obj['Line'] == 'D':
                    selected_ass_d.append(obj)
                if obj['Department'] == 'ASS' and obj['Line'] == 'E':
                    selected_ass_e.append(obj)
                if obj['Department'] == 'ASS' and obj['Line'] == 'F':
                    selected_ass_f.append(obj)
                if obj['Department'] == 'ASS' and obj['Line'] == 'J':
                    selected_ass_j.append(obj)

            else:
                message = 'Not found file ' + cur_path + '. Please check the path of file.'

    if request.method == 'POST':
        # Handle data Plan form
        if 'btn-plan' in request.form:
            data = {}
            log_name = 'Log.xlsx'
            log_writer = pd.ExcelWriter(path + log_name, engine='xlsxwriter')
            log_df_head = pd.DataFrame(
                columns=['Date', 'Time', 'Department', 'Line', 'Model', 'Plan', 'Version', 'File Name'])

            # Get data Plan form
            start = 0
            qty_actual = 0
            pause = 0
            time = datetime.now().strftime('%H:%M:%S')
            machine = 1
            material = 1
            quality = 1
            other = 1
            my_date = date.today().strftime('%Y-%m-%d')
            dept = request.form.get('department')
            line = request.form.get('line')
            model = request.form.get('model')
            version = request.form.get('version')
            qty_plan = int(request.form.get('plan'))
            per_plan = 0
            st_plan = 0
            for i in dict_model['Model']:
                if (dict_model['Model'][i] == model):
                    st_plan = dict_model[dept][i]
            st_actual = 0

            nonspace_model = model.replace(' ', '-')
            filename = str(my_date) + '_' + dept + '_' + line + '_' \
                                    + nonspace_model + '_' + str(qty_plan) \
                                    + '_ver' + version + '.xlsx'

            data['Start/Stop'] = start
            data['Q\'ty Actual'] = qty_actual
            data['Pause/Continue'] = pause
            data['Time'] = time
            data['Machine'] = machine
            data['Material'] = material
            data['Quality'] = quality
            data['Other'] = other
            data['Date'] = my_date
            data['Department'] = dept
            data['Line'] = line
            data['Model'] = model
            data['Version'] = version
            data['Q\'ty Plan'] = qty_plan
            data['% Plan'] = per_plan
            data['ST Plan'] = st_plan
            data['ST Actual'] = st_actual

            # print(data)

            # Check exist Log file
            if not os.path.isfile(path + log_name):
                # Create new Log file
                log_df_head.to_excel(
                    log_writer, sheet_name='Sheet1', index=False)
                log_writer.save()

            # Check exists Plan file
            location = dept + '/' + line + '/' + filename
            if os.path.isfile(path + location):
                # Show warning message
                message = '<b>You have a previous version of this plan. \
                    <br>Please change different version and try again.</b>'
                return render_template('demo.html', brand=brand, title='Demo App',
                                       data=dict_model, department=dict_dept,
                                       line=dict_line, msg=message, log=dict_log,
                                       selected_date=selected_date, selected_log=selected_log,
                                       selected_arm_a=selected_arm_a, selected_arm_b=selected_arm_b,
                                       selected_arm_c=selected_arm_c, selected_arm_d=selected_arm_d,
                                       selected_arm_e=selected_arm_e, selected_arm_f=selected_arm_f,
                                       selected_arm_j=selected_arm_j, selected_ass_j=selected_ass_j,
                                       selected_ass_a=selected_ass_a, selected_ass_b=selected_ass_b,
                                       selected_ass_c=selected_ass_c, selected_ass_d=selected_ass_d,
                                       selected_ass_e=selected_ass_e, selected_ass_f=selected_ass_f
                                       )
            else:
                # Create new Plan file
                message = ''
                writer = pd.ExcelWriter(path + location, engine='openpyxl')
                df = pd.DataFrame([data])
                df.to_excel(writer, sheet_name='Sheet1', index=False)
                writer.save()

                # Insert data to Log file
                wb = openpyxl.load_workbook(path+log_name)
                ws = wb.active
                new_row = ws.max_row + 1

                ws['A' + str(new_row)] = my_date
                ws['B' + str(new_row)] = time
                ws['C' + str(new_row)] = dept
                ws['D' + str(new_row)] = line
                ws['E' + str(new_row)] = model
                ws['F' + str(new_row)] = qty_plan
                ws['G' + str(new_row)] = version
                ws['H' + str(new_row)] = filename

                wb.save(path+log_name)

            return render_template('demo.html', brand=brand, title='Demo App',
                                   data=dict_model, department=dict_dept,
                                   line=dict_line, msg=message, log=dict_log,
                                   selected_date=selected_date, selected_log=selected_log,
                                   selected_arm_a=selected_arm_a, selected_arm_b=selected_arm_b,
                                   selected_arm_c=selected_arm_c, selected_arm_d=selected_arm_d,
                                   selected_arm_e=selected_arm_e, selected_arm_f=selected_arm_f,
                                   selected_arm_j=selected_arm_j, selected_ass_j=selected_ass_j,
                                   selected_ass_a=selected_ass_a, selected_ass_b=selected_ass_b,
                                   selected_ass_c=selected_ass_c, selected_ass_d=selected_ass_d,
                                   selected_ass_e=selected_ass_e, selected_ass_f=selected_ass_f
                                   )

        # Handle data History form
        if 'btn-history' in request.form:
            # Filter data
            selected_log.clear()  # remove data of current day
            selected_date = str(request.form['date-history'])

            selected_arm_a.clear()
            selected_arm_b.clear()
            selected_arm_c.clear()
            selected_arm_d.clear()
            selected_arm_e.clear()
            selected_arm_f.clear()
            selected_arm_j.clear()

            selected_ass_a.clear()
            selected_ass_b.clear()
            selected_ass_c.clear()
            selected_ass_d.clear()
            selected_ass_e.clear()
            selected_ass_f.clear()
            selected_ass_j.clear()

            # Get data of selected day in Log file
            for item in dict_log['Date']:
                if dict_log['Date'][item] == selected_date:
                    obj = {}
                    obj['Date'] = dict_log['Date'][item]
                    obj['Time'] = dict_log['Time'][item]
                    obj['Department'] = dict_log['Department'][item]
                    obj['Line'] = dict_log['Line'][item]
                    obj['Model'] = dict_log['Model'][item]
                    obj['Plan'] = dict_log['Plan'][item]
                    obj['Version'] = dict_log['Version'][item]
                    obj['File Name'] = dict_log['File Name'][item]
                    selected_log.append(obj)

            # Get data of selected day in Plan file
            for item in selected_log:
                # View history
                cur_path = path + 'ViewData/' + item['Department'] + '/' \
                                + item['Line'] + '/' \
                                + item['File Name']
                if os.path.isfile(cur_path):
                    plan_df = pd.read_excel(cur_path, 'Sheet1')
                    last_row = len(plan_df) + 1
                    obj = {}  # Object save last row data

                    # Open Plan file
                    wb = openpyxl.load_workbook(cur_path)
                    ws = wb.active

                    # Get data last line of Plan file
                    obj = {}
                    obj['Start/Stop'] = ws['A' + str(last_row)].value
                    obj['Q\'ty Actual'] = ws['B' + str(last_row)].value
                    obj['Pause/Continue'] = ws['C' + str(last_row)].value
                    obj['Time'] = ws['D' + str(last_row)].value
                    obj['Machine'] = ws['E' + str(last_row)].value
                    obj['Material'] = ws['F' + str(last_row)].value
                    obj['Quality'] = ws['G' + str(last_row)].value
                    obj['Other'] = ws['H' + str(last_row)].value
                    obj['Date'] = ws['I' + str(last_row)].value
                    obj['Department'] = ws['J' + str(last_row)].value
                    obj['Line'] = ws['K' + str(last_row)].value
                    obj['Model'] = ws['L' + str(last_row)].value
                    obj['Version'] = ws['M' + str(last_row)].value
                    obj['Q\'ty Plan'] = ws['N' + str(last_row)].value
                    obj['% Plan'] = round(ws['O' + str(last_row)].value, 1)
                    obj['ST Plan'] = round(ws['P' + str(last_row)].value, 2)
                    obj['ST Actual'] = round(ws['Q' + str(last_row)].value, 2)

                    wb.close()

                    if obj['Department'] == 'ARM' and obj['Line'] == 'A':
                        selected_arm_a.append(obj)
                    if obj['Department'] == 'ARM' and obj['Line'] == 'B':
                        selected_arm_b.append(obj)
                    if obj['Department'] == 'ARM' and obj['Line'] == 'C':
                        selected_arm_c.append(obj)
                    if obj['Department'] == 'ARM' and obj['Line'] == 'D':
                        selected_arm_d.append(obj)
                    if obj['Department'] == 'ARM' and obj['Line'] == 'E':
                        selected_arm_e.append(obj)
                    if obj['Department'] == 'ARM' and obj['Line'] == 'F':
                        selected_arm_f.append(obj)
                    if obj['Department'] == 'ARM' and obj['Line'] == 'J':
                        selected_arm_j.append(obj)

                    if obj['Department'] == 'ASS' and obj['Line'] == 'A':
                        selected_ass_a.append(obj)
                    if obj['Department'] == 'ASS' and obj['Line'] == 'B':
                        selected_ass_b.append(obj)
                    if obj['Department'] == 'ASS' and obj['Line'] == 'C':
                        selected_ass_c.append(obj)
                    if obj['Department'] == 'ASS' and obj['Line'] == 'D':
                        selected_ass_d.append(obj)
                    if obj['Department'] == 'ASS' and obj['Line'] == 'E':
                        selected_ass_e.append(obj)
                    if obj['Department'] == 'ASS' and obj['Line'] == 'F':
                        selected_ass_f.append(obj)
                    if obj['Department'] == 'ASS' and obj['Line'] == 'J':
                        selected_ass_j.append(obj)
                else:
                    message = 'Not found file ' + cur_path + '. Please check the path of file.'

            return render_template('demo.html', brand=brand, title='Demo App',
                                   data=dict_model, department=dict_dept,
                                   line=dict_line, msg=message, log=dict_log,
                                   selected_date=selected_date, selected_log=selected_log,
                                   selected_arm_a=selected_arm_a, selected_arm_b=selected_arm_b,
                                   selected_arm_c=selected_arm_c, selected_arm_d=selected_arm_d,
                                   selected_arm_e=selected_arm_e, selected_arm_f=selected_arm_f,
                                   selected_arm_j=selected_arm_j, selected_ass_j=selected_ass_j,
                                   selected_ass_a=selected_ass_a, selected_ass_b=selected_ass_b,
                                   selected_ass_c=selected_ass_c, selected_ass_d=selected_ass_d,
                                   selected_ass_e=selected_ass_e, selected_ass_f=selected_ass_f
                                   )

    return render_template('demo.html', brand=brand, title='Demo App',
                           data=dict_model, department=dict_dept,
                           line=dict_line, msg=message, log=dict_log,
                           selected_date=selected_date, selected_log=selected_log,
                           selected_arm_a=selected_arm_a, selected_arm_b=selected_arm_b,
                           selected_arm_c=selected_arm_c, selected_arm_d=selected_arm_d,
                           selected_arm_e=selected_arm_e, selected_arm_f=selected_arm_f,
                           selected_arm_j=selected_arm_j, selected_ass_j=selected_ass_j,
                           selected_ass_a=selected_ass_a, selected_ass_b=selected_ass_b,
                           selected_ass_c=selected_ass_c, selected_ass_d=selected_ass_d,
                           selected_ass_e=selected_ass_e, selected_ass_f=selected_ass_f
                           )
